import openpyxl
import categorygetter as cat
import charactergetter as char
import tracery
import random

def loadTripleList() -> list :
    triples = []
    triple = {}
    actions = []

    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\Data\\scealextrix\\Midpoints.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for row in sheet_obj.iter_rows(min_row = 2, max_col = 3, max_row = 2200):
        for cell in row:
            actions.append([x.strip() for x in cell.value.split(',')])
        triple["BMP"] = actions[0] 
        triple["MP"] = actions[1] 
        triple["AMP"] = actions[2] 
        triples.append(triple.copy())
        #print(triple)
        triple.clear()
        actions.clear()
    return triples

def getTriple(accion_inicial : str, triples : list) -> list:
    possible_triples = [tl for tl in triples if accion_inicial in tl["BMP"]]
    if (not possible_triples):
        return None
    return random.choice(possible_triples)

def getActionList( tl : list, point : str) -> list:
    if (point == "BMP"):
        return tl["BMP"]
    if (point == "MP"):
        return tl["MP"]
    if (point == "AMP"):
        return tl["AMP"]
    return None

def loadInitialBookends() -> list:
    action_pairs = []
    action_pair ={}
    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\Data\\scealextrix\\Initial bookend actions.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c1,c4 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row = 729),
                     sheet_obj.iter_rows(min_col = 4, max_col = 4, min_row = 2, max_row = 729)):
        action_pair["action"] = c1[0].value 
        action_pair["renderings"] = [x.strip() for x in c4[0].value.split(',')]
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs

def getInitialBookend(action_pairs : list, accion_inicial : str) -> str:
    for pair in action_pairs:
        if(pair["action"] == accion_inicial):
            return random.choice(pair["renderings"])
    return None

def loadClosingBookends() -> list:
    action_pairs = []
    action_pair ={}
    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\Data\\scealextrix\\Closing bookend actions.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c1,c4 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row = 244),
                     sheet_obj.iter_rows(min_col = 3, max_col = 3, min_row = 2, max_row = 244)):
        action_pair["action"] = c1[0].value 
        action_pair["renderings"] = [x.strip() for x in c4[0].value.split(',')]
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs

def getClosingBookend(action_pairs : list, accion_final : str) -> str:
    for pair in action_pairs:
        if  (pair["action"] == accion_final):
            return random.choice(pair["renderings"])
    return None




##################Main routine#########################
repetitions = set()
accion_inicial = "preach_to"
opening = "preach_to"
tripleta = "A" + accion_inicial + "B"
triples = loadTripleList()
sel_actions = set()

### podria empaquetar estos más... solo get Category u get char
categories = cat.loadCategoryList()
category = cat.getCategory(categories, accion_inicial)
characters = char.loadNocList() 
prog = char.getProg(characters, category)

A = " " + prog["Name"] +" "
B = " " + char.getOpp_by_simple(characters, prog)  + " "

##################Plot Creation#########################
for x in range(0,10):
    while (True):
        tp = getTriple(accion_inicial, triples)
        repetition_index = triples.index(tp)
        ##print(repetition_index) ##for debugging
        ##print(repetitions) 
        if (repetition_index not in repetitions or tp != None):
            break
##    if tp == None:
##       break
    repetitions.add(repetition_index)
    MPs = getActionList(tp, "MP")
    AMPs = getActionList(tp, "AMP")

    rules = {
        "origin": [",A#MP#B,A#AMP#B"],
        "MP" : MPs,
        "AMP" : AMPs
        }

    grammar = tracery.Grammar(rules)
    Mpamp = grammar.flatten("#origin#")
    tripleta += Mpamp
    accion_inicial = Mpamp.split(",")[2][1:-1]
    MP = Mpamp.split(",")[1][1:-1]
    sel_actions.add(accion_inicial)
    sel_actions.add(MP)


 
######################Printing############################

closing = tripleta.split(",")[-1].replace("A","").replace("B","")
opening_idiom = getInitialBookend(loadInitialBookends(), opening)
closing_idiom = getClosingBookend(loadClosingBookends(), closing)
plot = tripleta.replace("Apreach_toB", opening_idiom).replace(closing, closing_idiom).replace("A", A).replace("B", B)
print(plot)



