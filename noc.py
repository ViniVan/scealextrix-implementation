import openpyxl
import os

def loadTripleList() -> list :
    triples = []
    triple = {}
    actions = []

    path = os.path.join(os.path.dirname(__file__), "Midpoints.xlsx")
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for row in sheet_obj.iter_rows(min_row = 2, max_col = 3, max_row = 2200):
        for cell in row:
            actions.append([x.strip() for x in cell.value.split(',')])
        triple["BMP"] = actions[0] 
        triple["MP"] = actions[1] 
        triple["AMP"] = actions[2] 
        triples.append(triple.copy())
        #print(triple)
        triple.clear()
        actions.clear()
    return triples
triples = loadTripleList()

def loadInitialBookends() -> list:
    action_pairs = []
    action_pair ={}
    path = os.path.join(os.path.dirname(__file__),"Initial bookend actions.xlsx")
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c1,c4 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row = 729),
                     sheet_obj.iter_rows(min_col = 4, max_col = 4, min_row = 2, max_row = 729)):
        action_pair["action"] = c1[0].value 
        action_pair["renderings"] = [x.strip() for x in c4[0].value.split(',')]
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs
initial_bookends = loadInitialBookends()

def loadClosingBookends() -> list:
    action_pairs = []
    action_pair ={}
    path = os.path.join(os.path.dirname(__file__),"Closing bookend actions.xlsx")
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c1,c4 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row = 244),
                     sheet_obj.iter_rows(min_col = 3, max_col = 3, min_row = 2, max_row = 244)):
        action_pair["action"] = c1[0].value 
        action_pair["renderings"] = [x.strip() for x in c4[0].value.split(',')]
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs
closing_bookends = loadClosingBookends()

def loadIdioms() -> list:
    action_pairs = []
    action_pair ={}
    path = os.path.join(os.path.dirname(__file__),"Idiomatic actions.xlsx")
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c1,c5 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row = 820),
                     sheet_obj.iter_rows(min_col = 5, max_col = 5, min_row = 2, max_row = 820)):
        action_pair["action"] = c1[0].value 
        action_pair["renderings"] = [x.strip() for x in c5[0].value.split(',')]
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs
idioms = loadIdioms()

def loadConnectors() -> list:
    action_pairs = []
    action_pair ={}
    path = os.path.join(os.path.dirname(__file__),"Action pairs.xlsx")
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c2,c3,c4 in zip(sheet_obj.iter_rows(min_col = 2, max_col = 2, min_row = 2, max_row = 3699),
                     sheet_obj.iter_rows(min_col = 3, max_col = 3, min_row = 2, max_row = 3699),
                     sheet_obj.iter_rows(min_col = 4, max_col = 4, min_row = 2, max_row = 3699)):
        action_pair["pair"] = {'before' : c2[0].value, 'after': c4[0].value} 
        action_pair["link"] = c3[0].value
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs
connectors = loadConnectors() 