import openpyxl
import random
import os

def loadNocList() -> list:
    characters = []
    character = {}
   
    path = os.path.join(os.path.dirname(__file__),"NOC.xlsx")
    wb_obj = openpyxl.load_workbook(path, read_only= 'True' )#, data_only=True)
    sheet_obj = wb_obj.active
    max_rows = sheet_obj.max_row

    for c1, c8, c10, c15, c18, c21 ,c22, c23 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row =1031),
                                        sheet_obj.iter_rows(min_col = 8, max_col = 8, min_row = 2, max_row =1031),
                                        sheet_obj.iter_rows(min_col = 10, max_col = 10, min_row = 2, max_row =1031),
                                        sheet_obj.iter_rows(min_col = 15, max_col = 15, min_row = 2, max_row =1031),
                                        sheet_obj.iter_rows(min_col = 18, max_col = 18, min_row = 2, max_row =1031),
                                        sheet_obj.iter_rows(min_col = 21, max_col = 21, min_row = 2, max_row =1031),
                                        sheet_obj.iter_rows(min_col = 22, max_col = 22, min_row = 2, max_row =1031),
                                        sheet_obj.iter_rows(min_col = 23, max_col = 23, min_row = 2, max_row =1031)):
        character["Name"] = c1[0].value
        character["Politics"] = [x.strip() for x in c8[0].value.split(',')]   ## tengo que parsear la lista de nuevo
        character["Opponent"] = [x.strip() for x in c10[0].value.split(',') if x.strip() != "EMPTY"]  ##está en el personaje
        character["Domains"] = [x.strip() for x in c15[0].value.split(',')]   ## tengo que parsear la lista de nuevo
        character["Actor"] = [x.strip() for x in str(c18[0].value).split(',') if x.strip() != "EMPTY"]
        character["Group"] = [x.strip() for x in str(c21[0].value).split(',') if x.strip() != "EMPTY"] ## tengo que parsear la lista de nuevo
        character["World"] = [x.strip() for x in str(c22[0].value).split(',') if x.strip() != "EMPTY"] ## tengo que parsear la lista de nuevo
        character["Category"] = [x.strip() for x in c23[0].value.split(',')]
        characters.append(character.copy())
        character.clear()
    return characters

characters = loadNocList()

def getProg(category: str) -> dict:  
    possible_progs = [char for char in characters if category in char["Category"]]
    if (not possible_progs):
        return random.choice(characters)
    return random.choice(possible_progs)

##podría hacer una sola función y usar un argumento para escoger el tipo de extracción?

def getOpp_by_politics(prog : dict) -> str:
    if ("left" in prog["Politics"]):
        politics = "right"
    else:
        politics = "left"
    possible_opps = [opp for opp in characters if politics in opp["Politics"]]
    return random.choice(possible_opps)["Name"]

def getOpp_by_domain(prog : dict) -> str:
    possible_opps = [opp for opp in characters if any(domain in prog["Domains"] for domain in opp["Domains"])]
    return random.choice(possible_opps)["Name"]

def getOpp_by_world(prog : dict) -> str:      
    possible_opps = [opp for opp in characters if any(world in prog["World"] for world in opp["World"])]
    if (not possible_opps):
        return None
    return random.choice(possible_opps)["Name"]

def getOpp_by_simple(prog : dict) -> str:
    if (not prog["Opponent"]):
        return None
    return random.choice(prog["Opponent"])

def getOpp_by_group(prog:dict) -> str:
    possible_opps = [opp for opp in characters if any(group in prog["Group"] for group in opp["Group"])]
    if (not possible_opps):
        return None
    return random.choice(possible_opps)["Name"]

def getOpp_by_actor(prog:dict) -> str:
    possible_opps = [opp for opp in characters if (any(actor in prog["Actor"] for actor in opp["Actor"]) 
                                                   or any(name in prog["Name"] for name in opp["Actor"]))]
    if (not possible_opps):
        return None
    return random.choice(possible_opps)["Name"]
