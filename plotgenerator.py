import openpyxl
import categorygetter as cat
import charactergetter as char
import tracery
import random
import re

#cambio en change to sets
def loadTripleList() -> list :
    triples = []
    triple = {}
    actions = []

    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\Data\\scealextrix\\Midpoints.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for row in sheet_obj.iter_rows(min_row = 2, max_col = 3, max_row = 2200):
        for cell in row:
            actions.append([x.strip() for x in cell.value.split(',')])
        triple["BMP"] = actions[0] 
        triple["MP"] = actions[1] 
        triple["AMP"] = actions[2] 
        triples.append(triple.copy())
        #print(triple)
        triple.clear()
        actions.clear()
    return triples

def getTriple(accion_inicial : str, triples : list) -> list:
    possible_triples = [tl for tl in triples if accion_inicial in tl["BMP"]]
    if (not possible_triples):
        return None
    return random.choice(possible_triples)

def getActionList( tl : list, point : str) -> list:
    if (point == "BMP"):
        return tl["BMP"]
    if (point == "MP"):
        return tl["MP"]
    if (point == "AMP"):
        return tl["AMP"]
    return None

def loadInitialBookends() -> list:
    action_pairs = []
    action_pair ={}
    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\Data\\scealextrix\\Initial bookend actions.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c1,c4 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row = 729),
                     sheet_obj.iter_rows(min_col = 4, max_col = 4, min_row = 2, max_row = 729)):
        action_pair["action"] = c1[0].value 
        action_pair["renderings"] = [x.strip() for x in c4[0].value.split(',')]
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs

def getInitialBookend(accion_inicial : str) -> str:
    action_pairs = loadInitialBookends()
    for pair in action_pairs:
        if(pair["action"] == accion_inicial):
            return random.choice(pair["renderings"])
    return None

def loadClosingBookends() -> list:
    action_pairs = []
    action_pair ={}
    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\Data\\scealextrix\\Closing bookend actions.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c1,c4 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row = 244),
                     sheet_obj.iter_rows(min_col = 3, max_col = 3, min_row = 2, max_row = 244)):
        action_pair["action"] = c1[0].value 
        action_pair["renderings"] = [x.strip() for x in c4[0].value.split(',')]
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs

def getClosingBookend(accion_final : str) -> str:
    action_pairs = loadClosingBookends()
    for pair in action_pairs:
        if  (pair["action"] == accion_final):
            return random.choice(pair["renderings"])
    return None

def loadIdioms() -> list:
    action_pairs = []
    action_pair ={}
    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\Data\\scealextrix\\Idiomatic actions.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c1,c5 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2, max_row = 820),
                     sheet_obj.iter_rows(min_col = 5, max_col = 5, min_row = 2, max_row = 820)):
        action_pair["action"] = c1[0].value 
        action_pair["renderings"] = [x.strip() for x in c5[0].value.split(',')]
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs
action_pairs = loadIdioms()

def getIdiom(accion : str) -> str:
    for pair in action_pairs:
        if  (pair["action"] == accion):
            return random.choice(pair["renderings"])
    return None

def loadConnectors() -> list:
    action_pairs = []
    action_pair ={}
    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\Data\\scealextrix\\Action pairs.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only=True)#, data_only=True)
    sheet_obj = wb_obj.active

    for c2,c3,c4 in zip(sheet_obj.iter_rows(min_col = 2, max_col = 2, min_row = 2, max_row = 3699),
                     sheet_obj.iter_rows(min_col = 3, max_col = 3, min_row = 2, max_row = 3699),
                     sheet_obj.iter_rows(min_col = 4, max_col = 4, min_row = 2, max_row = 3699)):
        action_pair["pair"] = {'before' : c2[0].value, 'after': c4[0].value} 
        action_pair["link"] = c3[0].value
        action_pairs.append(action_pair.copy())
        action_pair.clear()
    return action_pairs




def main() :
    used_triples = set()
    sel_actions = set()
    triples = loadTripleList()
    connector_list = loadConnectors()
    accion_inicial = "are_escorted_by"
    opening = "are_escorted_by"
    plot = "A" + accion_inicial + "B"

    ################# Load Category and Characters ####################
    category = cat.getCategory(accion_inicial)
    protagonist= char.getProg(category)
    A = protagonist["Name"]
    B = char.getOpp_by_simple(protagonist)
    while (B == None):
        B = char.getOpp_by_simple(protagonist) ##aquí se puede quedar un ratote
        #hacer una elección top down desde lo más complicado hasta simple y si no dar un personaje al azar

    ##################Creación de Trama#########################
    as_beginning = True

    for x in range(0,10):
        while (True):
            tp = getTriple(accion_inicial, triples)
            if (tp == None):
                as_beginning = False
                break
            triple_index = triples.index(tp)
            if (triple_index not in used_triples):
                break
        if (as_beginning == False):
            break
        used_triples.add(triple_index)

        midpoints_list = getActionList(tp, "MP")
        afterpoints_list = getActionList(tp, "AMP")

        rules = {
            "origin": [",A#MP#B,A#AMP#B"],
            "MP" : midpoints_list,
            "AMP" : afterpoints_list
            }

        grammar = tracery.Grammar(rules)
        mid_after_points = grammar.flatten("#origin#")
        accion_inicial = mid_after_points.split(",")[2][1:-1] ## next first action
        midpoint = mid_after_points.split(",")[1][1:-1]
        sel_actions.add(accion_inicial)
        sel_actions.add(midpoint)
        plot += mid_after_points

        action_pair = {'before': accion_inicial, 'after': midpoint}
        for connector in connector_list:
            if (action_pair == connector['pair']):
                plot.replace(accion_inicial, accion_inicial + connector['link'])
        plot = plot.replace(",",". ")

    ######################Printing############################
    closing = accion_inicial
    opening_idiom = getInitialBookend(opening)
    closing_idiom = getClosingBookend(closing)
    if opening_idiom == None:
        opening_idiom = getIdiom(opening)
    if closing_idiom == None:
        closing_idiom = getIdiom(closing)

    plot = plot.replace("A"+opening+"B", opening_idiom).replace("A"+closing+"B", closing_idiom)

    for action in sel_actions:
        replacement = getIdiom(action)
        if (replacement != None):
            plot = plot.replace("A"+action+"B", replacement)

    print(plot.replace("A", A).replace("B", B))


main()



