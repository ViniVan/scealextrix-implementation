import openpyxl
import random

def loadNocList() -> list:
    characters = []
    character = {}
   
    path = "C:\\Users\\Vandré\\Documents\\UAM\Materias\\perez2\\PT\\DATA\\scealextrix\\NOC.xlsx"
    wb_obj = openpyxl.load_workbook(path, read_only= 'True' )#, data_only=True)
    sheet_obj = wb_obj.active
    max_rows = sheet_obj.max_row

    for c1, c8, c10, c15,c22, c23 in zip(sheet_obj.iter_rows(min_col = 1, max_col = 1, min_row = 2),
                                        sheet_obj.iter_rows(min_col = 8, max_col = 8, min_row = 2),
                                        sheet_obj.iter_rows(min_col = 10, max_col = 10, min_row = 2),
                                        sheet_obj.iter_rows(min_col = 15, max_col = 15, min_row = 2),
                                        sheet_obj.iter_rows(min_col = 22, max_col = 22, min_row = 2),
                                        sheet_obj.iter_rows(min_col = 23, max_col = 23, min_row = 2)):
        character["Name"] = c1[0].value
        character["Politics"] = [x.strip() for x in c8[0].value.split(',')]   ## tengo que parsear la lista de nuevo
        character["Opponent"] = [x.strip() for x in c10[0].value.split(',') if x.strip() != "EMPTY"]  ##está en el personaje
        character["Domains"] = [x.strip() for x in c15[0].value.split(',')]   ## tengo que parsear la lista de nuevo
        character["World"] = [x.strip() for x in str(c22[0].value).split(',') if x.strip() != "EMPTY"] ## tengo que parsear la lista de nuevo
        character["Category"] = [x.strip() for x in c23[0].value.split(',')]
        characters.append(character.copy())
        character.clear()
    return characters


def getProg(characters : list, category: str) -> dict:
    possible_progs = [char for char in characters if category in char["Category"]] 
    return random.choice(possible_progs)

def getOpp_by_politics(characters : list, prog : dict) -> str:
    if ("left" in prog["Politics"]):
        politics = "right"
    else:
        politics = "left"
    possible_opps = [opp for opp in characters if politics in opp["Politics"]]
    return random.choice(possible_opps)["Name"]

def getOpp_by_domain(characters:list, prog : dict) -> str:
    possible_opps = [opp for opp in characters if any(domain in prog["Domains"] for domain in opp["Domains"])]
    return random.choice(possible_opps)["Name"]

def getOpp_by_world(characters:list, prog : dict) -> str:      #aquí hay un problemas con las celdas EMpTY
    possible_opps = [opp for opp in characters if any(world in prog["World"] for world in opp["World"])]
    if (not possible_opps):
        return None
    return random.choice(possible_opps)["Name"]

def getOpp_by_simple(characters : list, prog : dict) -> str:  
    return random.choice(prog["Opponent"])
    
##    
##chars = loadNocList()
###print(chars)
##prog = getProg(chars, "Economist")
###opp = getOpp_by_politics(chars, prog)
###opp = getOpp_by_domain(chars,prog)
##opp = getOpp_by_world(chars, prog)
###opp = getOpp_by_simple(chars, prog)
##print(prog["Name"])
##print(opp)
